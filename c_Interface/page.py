from .modules import resources
from .modules import *
from PySide6.QtWidgets import QApplication, QWidget, QPushButton, QLineEdit, QVBoxLayout, QFileDialog, QHeaderView, QMessageBox
import sys
import os
import pandas as pd
import re
from d_function import d_snapshot, a_topology, c_xueshi, b_visible, e_interact
import openpyxl
import networkx as nx
import matplotlib.pyplot as plt
from PySide6.QtGui import QPixmap
from d_function.e_interact import window

class MainWindow(QMainWindow):
    ##############################################################
    def __init__(self):
        # 调用父类的构造函数以初始化QMainWindow
        QMainWindow.__init__(self)

        # 隐藏原本的菜单栏
        self.setWindowFlags(Qt.FramelessWindowHint)
        self.setAttribute(Qt.WA_TranslucentBackground)

        # 将小部件设置为全局
        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)
        global widgets
        widgets = self.ui

        # 添加阴影效果
        self.shadow = QGraphicsDropShadowEffect(self)
        self.shadow.setBlurRadius(17)
        self.shadow.setXOffset(0)
        self.shadow.setYOffset(0)
        self.shadow.setColor(QColor(0, 0, 0, 150))
        self.ui.classApp.setGraphicsEffect(self.shadow)

        # 设置主页并选择菜单
        widgets.myStackedWidget.setCurrentWidget(widgets.page_1)

        # 初始化数据结构G、subsets
        self.G = nx.DiGraph()
        self.subsets = []

        # 标记现在主题为亮色
        self.useCustomTheme = True
        ########################################################
        # 表格均设置为列宽一致
        widgets.page1TableWidget.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        widgets.page2TableWidget.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        widgets.page4TableWidget.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)

        # 窗口按钮点击事件
        widgets.minimizeAppBtn.clicked.connect(self.showMinimized)
        widgets.maximizeRestoreAppBtn.clicked.connect(self.toggleMaximizeRestore)
        widgets.closeAppBtn.clicked.connect(self.close)
        widgets.colorTopBtn.clicked.connect(self.buttonClick)

        # 具体任务按钮点击事件
        widgets.btn_home.clicked.connect(self.buttonClick)
        widgets.btn_graphic.clicked.connect(self.buttonClick)
        widgets.btn_human.clicked.connect(self.buttonClick)
        widgets.btn_look.clicked.connect(self.buttonClick)
        widgets.page1ChooseFileBtn.clicked.connect(self.buttonClick) #打开文件
        widgets.page4PngBtn.clicked.connect(self.buttonClick)
        widgets.page4LeftBtn.clicked.connect(self.buttonClick)
        widgets.page4RightBtn.clicked.connect(self.buttonClick)
        widgets.page4XueShiBtn.clicked.connect(self.buttonClick)
        widgets.clearPhotoBtn.clicked.connect(self.buttonClick)
        widgets.reloadPhotoBtn.clicked.connect(self.buttonClick)

        # 按钮点击高亮事件
        widgets.page4RightBtn.clicked.connect(self.courseSearch)
        widgets.page4LeftBtn.clicked.connect(self.courseSearch)

        # 文本框内容改变事件
        widgets.page1ChooseClassLineEdit.textChanged.connect(self.classSearch)
        widgets.page4AdjustLinetext.textChanged.connect(self.courseSearch)
        ########################################################
        # 显示应用程序
        self.show()
    ##############################################################

    ##############################################################
    def mousePressEvent(self, event):
        if event.button() == Qt.LeftButton:
            self.start_pos = event.globalPos()
            self.dragging = True
    def mouseMoveEvent(self, event):
        if self.dragging:
            delta = event.globalPos() - self.start_pos
            self.move(self.mapToGlobal(delta))
            self.start_pos = event.globalPos()
    def mouseReleaseEvent(self, event):
        self.dragging = False
    ##############################################################


    ##############################################################
    # 编写最大化/还原按钮的点击事件处理函数
    def toggleMaximizeRestore(self):
        if self.isMaximized():
            self.showNormal()
            self.ui.maximizeRestoreAppBtn.setIcon(QIcon(u":/icons/images/icons/icon_maximize.png"))
            window.set_geometry(720,520,1200,600)
            window.show()
        else:
            self.showMaximized()
            self.ui.maximizeRestoreAppBtn.setIcon(QIcon(u":/icons/images/icons/icon_restore.png"))
            window.set_geometry(150,200,2350,1200)
            window.show()
    ##############################################################


    ##############################################################
    def classSearch(self):
        search_class = widgets.page1ChooseClassLineEdit.text()
        if search_class:
            for row in range(widgets.page1TableWidget.rowCount()):
                item = widgets.page1TableWidget.item(row, 0)  # 要查找的列是第一列
                if item and item.text() == search_class:
                    # 找到匹配的值，将该行高亮显示
                    widgets.page1TableWidget.selectRow(row)
                    widgets.page1TableWidget.selectedItems()
                    break  # 找到后退出循环
        else:
            # 如果搜索值为空，取消高亮显示
            widgets.page1TableWidget.clearSelection()
    def courseSearch(self):
        search_class = widgets.page4AdjustLinetext.text()
        successors = a_topology.find_node_and_successors(self.G, search_class)
        if search_class:
            for row in range(widgets.page4TableWidget.rowCount()):
                for col in range(widgets.page4TableWidget.columnCount()):
                    item = widgets.page4TableWidget.item(row, col)  # 要查找的列是第一列
                    if item and (item.text() == search_class or item.text() in successors):
                        # 找到匹配的值，将该单元格高亮显示
                        widgets.page4TableWidget.item(row, col).setSelected(True)
        else:
            # 如果搜索值为空，取消高亮显示
            widgets.page4TableWidget.clearSelection()
    ##############################################################


    ############################################################## 功能性函数：读数据部分
    def clearContent(self):
        widgets.page1TableWidget.clearContents()
        widgets.page2TableWidget.clearContents()
        widgets.page4TableWidget.clearContents()
        widgets.page1ChooseClassLineEdit.clear()
        widgets.label.clear()
        widgets.page4AdjustLinetext.clear()
        widgets.page4XueShiLinetext.clear()
    def readXlsxFile(self, fp):
        if not fp or not os.path.isfile(fp):
            QMessageBox.warning(self, "警告", "输入文件有误！", QMessageBox.Ok)
            return None

        excel1 = openpyxl.load_workbook(fp)
        sheet1 = excel1.active
        df = pd.DataFrame(sheet1.values, columns=["课程名称", "学分", "总学时", "先修课程", "简介"])
        excel1.close()
        return df
    def writeInTable1(self, dataform):
        num_rows, _ = dataform.shape
        if widgets.page1TableWidget.rowCount() < num_rows:
            widgets.page1TableWidget.setRowCount(num_rows)
        # 循环读取DataFrame行，并插入到page1TableWidget中
        for row_index, (_, row) in enumerate(dataform.iterrows(), start=0):
            for col_index, value in enumerate(row, start=0):
                item = QTableWidgetItem(str(value))
                widgets.page1TableWidget.setItem(row_index - 1, col_index, item)
    def writeInTable2(self, dataform):
        pair = [] # 存储先修课程和后修课程
        for index, row in dataform.iterrows():
            course_name = row["课程名称"]
            prerequisite_course_text = row["先修课程"]
            # 使用正则表达式来分割先修课程，支持逗号、顿号、空格或者分号分隔
            prerequisite_course_list = re.split(r'[，、；, \s]+', prerequisite_course_text) if pd.notna(prerequisite_course_text) else []
            # 分有无前置课程两种情况
            if not prerequisite_course_list:
                pair.append(("无", course_name))
            else:
                for prereq in prerequisite_course_list:
                    pair.append((prereq.strip(), course_name))

        # 行数
        num_rows = len(pair)
        widgets.page2TableWidget.setRowCount(num_rows)

        # 遍历写入
        for row_index, (prereq, course) in enumerate(pair):
            item_prereq = QTableWidgetItem(prereq)
            widgets.page2TableWidget.setItem(row_index - 1, 0, item_prereq)  # 先修课程第一列
            item_course = QTableWidgetItem(course)
            widgets.page2TableWidget.setItem(row_index - 1, 1, item_course)  # 后修课程第二列
    def getStuTime(self, dataform):
        global data_dict
        data_dict = {}  # 学时字典
        for index, row in dataform.iterrows():
            course_name = row["课程名称"]
            student_time = row["总学时"]
            data_dict[course_name] = student_time
    def saveTable2ToXlsx(self):
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        # 遍历 page2TableWidget 中的行和列，并将数据写入.xlsx工作表
        for row_index in range(widgets.page2TableWidget.rowCount()):
            for col_index in range(widgets.page2TableWidget.columnCount()):
                item = widgets.page2TableWidget.item(row_index, col_index)
                if item:
                    cell_value = item.text()
                    worksheet.cell(row=row_index + 1, column=col_index + 1, value=cell_value)
        #workbook.save('b_data/final_data/信安课程先修后修.xlsx')
        workbook.save('b_data/final_data/计科课程先修后修.xlsx')
    ##############################################################

    # 按钮点击事件
    def buttonClick(self):
        # 获取被点击的按钮
        btn = self.sender()
        btnName = btn.objectName()
        ##############################################################
        if btnName == "btn_home":
            widgets.myStackedWidget.setCurrentWidget(widgets.page_1)
        if btnName == "btn_graphic":
            widgets.myStackedWidget.setCurrentWidget(widgets.page_2)
        if btnName == "btn_human":
            widgets.myStackedWidget.setCurrentWidget(widgets.page_3)
            e_interact.run_application(720,520,1200,600)
        if btnName == "btn_look":
            widgets.myStackedWidget.setCurrentWidget(widgets.page_4)
        ##############################################################

        ##############################################################
        if btnName == "colorTopBtn":
            if self.useCustomTheme:
                # 在按钮点击时，修改按钮的样式
                str1 = open("c_Interface/themes/py_dracula_dark.qss", 'r').read()
                self.ui.styleSheet.setStyleSheet(str1)
                self.useCustomTheme = False
            else:
                str1 = open("c_Interface/themes/py_dracula_light.qss", 'r').read()
                self.ui.styleSheet.setStyleSheet(str1)
                self.useCustomTheme = True
        ##############################################################

        ##############################################################
        if btnName == "page1ChooseFileBtn":
            self.clearContent()
            ######################## 得到文件
            file_dialog = QFileDialog(self)
            file_dialog.setNameFilter("Excel 文件 (*.xlsx)")
            file_dialog.setFileMode(QFileDialog.ExistingFile)
            file_dialog.setViewMode(QFileDialog.List)
            ########################  写入文本框
            if file_dialog.exec():
                selected_files = file_dialog.selectedFiles()
                if selected_files:  widgets.page1ChooseFileLineEdit.setText(selected_files[0])
            ########################  得到数据
            if widgets.page1ChooseFileLineEdit.text() != "":
                file_path = widgets.page1ChooseFileLineEdit.text()
                df = self.readXlsxFile(file_path)
                ########################## 写入表1、2，得到学时，保存先修后修关系
                self.writeInTable1(df)
                self.writeInTable2(df)
                self.getStuTime(df)
                self.saveTable2ToXlsx()
                ########################  产生图，拓扑排序后，写入表4
                # self.G = a_topology.generate_Graph("b_data/final_data/信安课程先修后修.xlsx")
                self.G = a_topology.generate_Graph("b_data/final_data/计科课程先修后修.xlsx")
                self.G, self.subsets = a_topology.topo_sort(self.G)
                widgets.page4TableWidget = a_topology.add_to_table(widgets.page4TableWidget, self.subsets)
                ########################  写入label图片pixmap
                # 创建TopologicalSortGraph对象并绘制图形
                # topo_sort_graph = b_visible.TopologicalSortGraph(self.G)
                # topo_sort_graph.plot_topological_sort()
                # pixmap = QPixmap("b_data/final_data/信安初始课程拓扑.png")
                pixmap = QPixmap("b_data/final_data/计科初始课程拓扑.png")
                widgets.label.setScaledContents(True)  # 缩放内容以适应 QLabel
                widgets.label.setPixmap(pixmap)
        ##############################################################

        ##############################################################
        if btnName == "page4PngBtn":
            d_snapshot.take_snapshot(widgets.page4TableWidget)
        ##############################################################

        ##############################################################
        if btnName == "page4LeftBtn":
            isok = False
            adjust_class = widgets.page4AdjustLinetext.text()
            for row in range(widgets.page4TableWidget.rowCount()):
                for col in range(widgets.page4TableWidget.columnCount()):
                    item = widgets.page4TableWidget.item(row, col)
                    if item and item.text() == adjust_class:
                        results = a_topology.find_node_and_predecessors(self.G, adjust_class)
                        for result in results:
                            if "无意义" in result:  # 可以向左移动
                                isok = True
                                self.G = a_topology.delete_virtual_node(self, self.G, adjust_class)
                                self.G, self.subsets = a_topology.topo_sort(self.G)
                                widgets.page4TableWidget.clearContents()
                                widgets.page4TableWidget = a_topology.add_to_table(widgets.page4TableWidget, self.subsets)
                                break
                    if isok:
                        break  # 添加 break，终止外层循环
            if not isok:
                QMessageBox.warning(self, "警告", "该移动课程已在其允许的最前一列！", QMessageBox.Ok)
        ##############################################################

        ##############################################################
        if btnName == "page4RightBtn":
            isok = False
            warning_shown = False
            adjust_class = widgets.page4AdjustLinetext.text()
            # 遍历表格的所有行和列
            for row in range(widgets.page4TableWidget.rowCount()):
                for col in range(widgets.page4TableWidget.columnCount()):
                    item = widgets.page4TableWidget.item(row, col)
                    if item and item.text() == adjust_class:
                        for row_index in range(widgets.page4TableWidget.rowCount()):
                            item_at_eighth_col = widgets.page4TableWidget.item(row_index, 7)  # 第八列
                            if item_at_eighth_col is None:
                                break
                            if item_at_eighth_col.text() in a_topology.find_node_and_successors(self.G, adjust_class):
                                isok = True
                                break  # 添加 break，表示已经找到符合条件的情况，不再继续检查当前行的其它列
                        break  # 添加 break，表示已经找到符合条件的情况，不再继续检查其它行
            if isok:
                if not warning_shown:
                    QMessageBox.warning(self, "警告", "其本身或后置课程已在最后一学期，将超出课表容纳范围！",
                                        QMessageBox.Ok)
                    warning_shown = True
            else:
                # 如果调整的课程和其后续课程不在第八列，执行以下操作
                self.G = a_topology.add_virtual_node(self.G, adjust_class)
                self.G, self.subsets = a_topology.topo_sort(self.G)
                widgets.page4TableWidget.clearContents()
                widgets.page4TableWidget = a_topology.add_to_table(widgets.page4TableWidget, self.subsets)
        ############################################################

        ##############################################################
        if btnName == "page4XueShiBtn":
            all_xueshi = int(widgets.page4XueShiLinetext.text())
            # # 打印 xueshi_judge 调用前的 self.G
            # print("Before xueshi_judge, self.G:", self.G.nodes, self.G.edges)
            G_copy = self.G.copy()
            sbsets = c_xueshi.xueshi_judge(all_xueshi, G_copy, data_dict)
            # # 打印 xueshi_judge 调用后的 self.G
            # print("After xueshi_judge, self.G:", self.G.nodes, self.G.edges)
            widgets.page4TableWidget.clearContents()
            widgets.page4TableWidget = a_topology.add_to_table(widgets.page4TableWidget, sbsets)
        ############################################################

        ##############################################################
        if btnName == "clearPhotoBtn":
            widgets.label.setScaledContents(False)  # 不要缩放内容（丑）
            widgets.label.setPixmap(QtGui.QPixmap(":/images/images/images/PyDracula_vertical.png"))
        ############################################################

        ##############################################################
        if btnName == "reloadPhotoBtn":
            widgets.label.setScaledContents(True)
            topo_sort_graph = b_visible.TopologicalSortGraph(self.G)
            topo_sort_graph.plot_topological_sort()
            # widgets.label.setPixmap(QtGui.QPixmap("b_data/final_data/信安专业课程拓扑.png"))
            widgets.label.setPixmap(QtGui.QPixmap("b_data/final_data/计科专业课程拓扑.png"))
        ##############################################################